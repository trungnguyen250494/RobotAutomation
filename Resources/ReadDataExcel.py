import openpyxl

vk = openpyxl.load_workbook("../TestData/test.xlsx")


def fetch_number_of_rows(SheetName):
    sh = vk[SheetName]
    return sh.max_row


def fetch_cell_data(SheetName, row, column):
    sh = vk[SheetName]
    cell = sh.cell(int(row), int(column))
    return cell.value


print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1", 1, 1))
